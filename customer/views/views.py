import json
from datetime import datetime

import openpyxl
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.shortcuts import render,  redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.views import View
from django.views.generic import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin
from openpyxl.utils import get_column_letter

from customer.forms import CustomerModelForm
from customer.models import Customer
import csv


# Function-based :
# def customer_list(request):
#     search = request.GET.get('search')
#     filter_button_clicked = request.GET.get('filter')
#     customers = Customer.objects.raw('''SELECT * FROM customer_customer;''')
#     for customer in customers:
#         print(customer.email)
#     if filter_button_clicked:
#         customers = customers.order_by('-updated_at')
#     if search:
#         customers = customers.filter(
#             Q(first_name__icontains=search) |
#             Q(last_name__icontains=search) |
#             Q(email__icontains=search) |
#             Q(phone__icontains=search) |
#             Q(address__icontains=search)
#         )
#     paginator = Paginator(customers, 4)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)
#     return render(request, 'customer/customer-list.html', {
#         'customers': page_obj,
#         'filter_button_clicked': filter_button_clicked
#     })

# def customer_details(request, customer_slug):
#     customer = get_object_or_404(Customer, slug=customer_slug)
#     context = {'customer': customer}
#     return render(request, 'customer/customer-details.html', context)

# @login_required
# def new_customer(request):
#     form = CustomerModelForm()
#     if request.method == 'POST':
#         form = CustomerModelForm(request.POST, request.FILES)
#         if form.is_valid():
#             form.save()
#             return redirect('customer_list')
#     context = {'form': form}
#     return render(request, 'customer/new_customer.html', context)

# @login_required
# def export_customers(request):
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="customers.csv"'
#     writer = csv.writer(response)
#     writer.writerow(['ID', 'Full_Name', 'Email', 'Updated At'])
#     for customer in Customer.objects.all():
#         writer.writerow([customer.id, customer.full_name, customer.email, customer.updated_at])
#     return response

# @login_required
# def profile_page(request):
#     if request.method == 'POST':
#         user = User.objects.get(username=request.POST['username'])
#         context = {'user': user}
#         return render(request, 'customer/profile_page.html', context)
#     return render(request, 'customer/profile_page.html')

# @login_required
# def delete_customer(request, customer_slug):
#     customer = get_object_or_404(Customer, slug=customer_slug)
#     customer.delete()
#     return redirect('customer_list')

# @login_required
# def edit_customer(request, customer_slug):
#     customer = get_object_or_404(Customer, slug=customer_slug)
#     form = CustomerModelForm(instance=customer)
#     if request.method == 'POST':
#         form = CustomerModelForm(request.POST, request.FILES, instance=customer)
#         if form.is_valid():
#             form.save()
#             return redirect('customer_list')
#     return render(request, 'customer/edit_customer.html', {'customer': customer, 'form': form})

# Class-based views:


class CustomerListView(View):

    def get(self, request):
        customers = Customer.objects.all()
        paginator = Paginator(customers, 4)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context = {'customers': page_obj}
        return render(request, 'customer/customer-list.html', context)

    # @login_required
    # def export_customers(request):
    #     response = HttpResponse(content_type='text/csv')
    #     response['Content-Disposition'] = 'attachment; filename="customers.csv"'
    #     writer = csv.writer(response)
    #     writer.writerow(['ID', 'Full_Name', 'Email', 'Updated At'])
    #     for customer in Customer.objects.all():
    #         writer.writerow([customer.id, customer.full_name, customer.email, customer.updated_at])
    #     return response

    def get_queryset(self):
        search = self.request.GET.get('search')
        filter_button_clicked = self.request.GET.get('filter')
        queryset = Customer.objects.all()

        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search) |
                Q(address__icontains=search)
            )

        if filter_button_clicked:
            queryset = queryset.order_by('-updated_at')

        return queryset


class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'customer/customer-details.html'
    context_object_name = 'customer'
    slug_field = 'slug'
    slug_url_kwarg = 'customer_slug'

class CustomerCreateView(LoginRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerModelForm
    template_name = 'customer/new_customer.html'
    success_url = '/customer-list/'

class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    model = Customer
    form_class = CustomerModelForm
    template_name = 'customer/edit_customer.html'
    slug_field = 'slug'
    slug_url_kwarg = 'customer_slug'
    success_url = '/customer-list/'

class CustomerDeleteView(LoginRequiredMixin, DeleteView):
    model = Customer
    template_name = 'customer/delete_customer.html'
    slug_field = 'slug'
    slug_url_kwarg = 'customer_slug'
    success_url = 'customer_list'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return redirect(self.success_url)

class CustomerExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="customers.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Full_Name', 'Email', 'Updated At'])
        for customer in Customer.objects.all():
            writer.writerow([customer.id, customer.full_name, customer.email, customer.updated_at])
        return response

class UserProfileView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        return render(request, 'customer/profile_page.html')

    def post(self, request, *args, **kwargs):
        user = User.objects.get(username=request.POST['username'])
        context = {'user': user}
        return render(request, 'customer/profile_page.html', context)

class ShoppingCartView(View):
    def get(self, request, *args, **kwargs):
        return render(request, 'shop/shopping-cart.html')

class CheckoutView(View):
    def get(self, request, *args, **kwargs):
        return render(request, 'shop/checkout.html')

    def post(self, request, *args, **kwargs):
        return render(request, 'shop/checkout.html')

class ExportDataView(View):
    def get(self, request, *args, **kwargs):
        date = datetime.now().strftime("%Y-%m-%d")
        export_format = request.GET.get('format')
        if export_format == 'csv':
            return self.export_csv(date)
        elif export_format == 'json':
            return self.export_json(date)
        elif export_format == 'xlsx':
            return self.export_xlsx(date)
        else:
            return HttpResponse(status=404, content='Bad Request')

    def export_csv(self, date):
        meta = Customer._meta
        field_names = [field.name for field in meta.fields]
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={Customer._meta.object_name}-{date}.csv'

        writer = csv.writer(response)
        writer.writerow(field_names)
        for obj in Customer.objects.all():
            writer.writerow([getattr(obj, field) for field in field_names])

        return response

    def export_json(self, date):
        data = list(Customer.objects.all().values('id', 'first_name', 'last_name', 'phone', 'email'))
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename={Customer._meta.object_name}-{date}.json'
        response.write(json.dumps(data, indent=4))
        return response

    def export_xlsx(self, date):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={Customer._meta.object_name}-{date}.xlsx'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Customers'

        # Field names
        fields = ['ID', 'First Name', 'Last Name', 'Phone', 'Email']
        for col_num, field_name in enumerate(fields, 1):
            column_letter = get_column_letter(col_num)
            worksheet[f'{column_letter}1'] = field_name

        # Data rows
        for row_num, obj in enumerate(Customer.objects.all(), 2):
            worksheet[f'A{row_num}'] = obj.id
            worksheet[f'B{row_num}'] = obj.first_name
            worksheet[f'C{row_num}'] = obj.last_name
            worksheet[f'D{row_num}'] = obj.phone
            worksheet[f'E{row_num}'] = obj.email

        workbook.save(response)
        return response